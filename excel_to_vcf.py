import streamlit as st
import pandas as pd
import re
import io
import zipfile
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

st.set_page_config(page_title="엑셀 → 매니저별 연락처 변환기", page_icon="📇", layout="centered")

# ── 스타일 ──
st.markdown("""
<style>
    .stApp { background-color: #f8fafc; }
    h1, h2, h3 { color: #1e293b !important; }
    .step-badge {
        display: inline-flex; align-items: center; justify-content: center;
        background: #2563eb; color: #fff; width: 26px; height: 26px;
        border-radius: 50%; font-size: 13px; font-weight: 700; margin-right: 8px;
    }
    .count-badge {
        background: #dbeafe; color: #2563eb;
        padding: 4px 14px; border-radius: 20px; font-weight: 600; font-size: 14px;
    }
    .success-box {
        background: #f0fdf4; border: 1px solid #bbf7d0;
        border-radius: 12px; padding: 16px 20px; color: #16a34a; margin-top: 12px;
    }
    .guide-box {
        background: #f1f5f9; border: 1px solid #e2e8f0;
        border-radius: 12px; padding: 18px 22px; color: #475569; font-size: 14px; line-height: 2;
    }
</style>
""", unsafe_allow_html=True)


# ── 함수 ──
def normalize_phone(raw):
    """전화번호 정규화"""
    if raw is None or (isinstance(raw, float) and pd.isna(raw)):
        return ""
    s = re.sub(r"[^0-9+]", "", str(raw).strip())
    if not s:
        return ""
    if s.startswith("+82"):
        s = "0" + s[3:]
    elif s.startswith("82") and len(s) > 10:
        s = "0" + s[2:]
    if not s.startswith("0") and len(s) >= 9:
        s = "0" + s
    return s


def vcf_escape(s):
    s = str(s)
    s = s.replace("\\", "\\\\")
    s = s.replace("\n", "\\n").replace("\r", "")
    s = s.replace(",", "\\,").replace(";", "\\;")
    return s


def to_vcard(name, phone, org="", note=""):
    clean_phone = normalize_phone(phone)
    lines = [
        "BEGIN:VCARD",
        "VERSION:3.0",
        f"FN:{vcf_escape(name)}",
        f"N:;{vcf_escape(name)};;;",
    ]
    if clean_phone:
        lines.append(f"TEL;TYPE=CELL:{clean_phone}")
    if org:
        lines.append(f"ORG:{vcf_escape(org)}")
    if note:
        lines.append(f"NOTE:{vcf_escape(note)}")
    lines.append("END:VCARD")
    return "\r\n".join(lines) + "\r\n"


def sanitize_filename(name):
    name = re.sub(r'[<>:"/\\|?*\r\n\t]', "_", str(name))
    return name.strip().rstrip(".") or "_"


def cell_value(v):
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    return str(v).strip()


def col_label(letter, header_preview):
    if header_preview:
        return f"{letter} — {header_preview}"
    return letter


# ── UI ──
st.markdown("## 📇 엑셀 → 매니저별 연락처 변환기")
st.caption("엑셀 **열(A, B, C...)을 직접 선택**하여 지점/매니저별 VCF로 변환합니다")

# Step 1: 파일 업로드
st.markdown('<span class="step-badge">1</span> **엑셀 파일 업로드**', unsafe_allow_html=True)

uploaded = st.file_uploader(
    "파일 선택 (.xlsx, .xls)",
    type=["xlsx", "xls"],
    label_visibility="collapsed",
)

if uploaded:
    try:
        wb = load_workbook(uploaded, data_only=True, read_only=False)
    except Exception as e:
        st.error(f"파일을 읽을 수 없습니다: {e}")
        st.stop()

    sheet_names = wb.sheetnames
    if len(sheet_names) > 1:
        sheet_name = st.selectbox("시트 선택", options=sheet_names, index=0)
    else:
        sheet_name = sheet_names[0]
    ws = wb[sheet_name]

    max_col = ws.max_column
    max_row = ws.max_row
    st.caption(f"시트: **{sheet_name}** · 열: {max_col}개 · 행: {max_row}행")

    # Step 2: 헤더/데이터 시작 행
    st.markdown("---")
    st.markdown('<span class="step-badge">2</span> **헤더 행 / 데이터 시작 행**', unsafe_allow_html=True)

    c1, c2 = st.columns(2)
    with c1:
        header_row = st.number_input(
            "헤더로 보여줄 행 (1부터)",
            min_value=1, max_value=min(20, max_row), value=min(2, max_row), step=1,
        )
    with c2:
        data_start_row = st.number_input(
            "데이터 시작 행 (1부터)",
            min_value=1, max_value=max_row, value=min(int(header_row) + 1, max_row), step=1,
        )

    # 헤더 미리보기
    header_preview = {}
    header_cells = list(ws.iter_rows(min_row=int(header_row), max_row=int(header_row), values_only=True))[0]
    for idx, val in enumerate(header_cells, start=1):
        letter = get_column_letter(idx)
        header_preview[letter] = cell_value(val)

    all_letters = [get_column_letter(i) for i in range(1, max_col + 1)]
    col_options = ["— 선택 안함 —"] + [col_label(l, header_preview.get(l, "")) for l in all_letters]
    letter_for_option = {col_label(l, header_preview.get(l, "")): l for l in all_letters}

    def pick_default(keywords):
        for l in all_letters:
            h = header_preview.get(l, "")
            for kw in keywords:
                if kw and kw in h:
                    return col_label(l, h)
        return None

    BRANCH_KW = ["지점"]
    MANAGER_KW = ["매니저"]
    AGENCY_KW = ["대리점"]
    SUB_KW = ["지사"]
    NAME_KW = ["이름", "팀장", "성명"]
    PHONE_KW = ["휴대전화", "휴대폰", "연락처", "전화"]
    NOTE_KW = ["메모", "비고", "직책"]

    # Step 3: 열 매핑
    st.markdown("---")
    st.markdown('<span class="step-badge">3</span> **열(A, B, C...) 매핑**', unsafe_allow_html=True)
    st.caption("실제 엑셀 열 문자를 골라주세요. 헤더 미리보기가 함께 표시됩니다.")

    def select_col(label, default_keywords, required=False):
        default = pick_default(default_keywords)
        idx = col_options.index(default) if default in col_options else 0
        suf = " *" if required else ""
        return st.selectbox(label + suf, options=col_options, index=idx, key=f"sel_{label}")

    cA, cB = st.columns(2)
    with cA:
        branch_sel = select_col("지점", BRANCH_KW, required=True)
        agency_sel = select_col("대리점", AGENCY_KW)
        name_sel = select_col("이름(팀장명)", NAME_KW, required=True)
        note_sel = select_col("메모", NOTE_KW)
    with cB:
        manager_sel = select_col("매니저", MANAGER_KW, required=True)
        sub_sel = select_col("지사", SUB_KW)
        phone_sel = select_col("휴대폰", PHONE_KW, required=True)

    def to_letter(sel):
        return letter_for_option.get(sel) if sel != "— 선택 안함 —" else None

    L_BRANCH = to_letter(branch_sel)
    L_MANAGER = to_letter(manager_sel)
    L_AGENCY = to_letter(agency_sel)
    L_SUB = to_letter(sub_sel)
    L_NAME = to_letter(name_sel)
    L_PHONE = to_letter(phone_sel)
    L_NOTE = to_letter(note_sel)

    if not (L_BRANCH and L_MANAGER and L_NAME and L_PHONE):
        st.warning("지점·매니저·이름·휴대폰 컬럼은 필수입니다.")
        st.stop()

    # Step 4: 추가 옵션
    st.markdown("---")
    st.markdown('<span class="step-badge">4</span> **추가 옵션**', unsafe_allow_html=True)

    oc1, oc2, oc3 = st.columns(3)
    with oc1:
        extra_note = st.text_input("메모에 추가로 붙일 텍스트 (선택)", value="")
    with oc2:
        name_prefix = st.text_input(
            "이름 앞에 붙일 문자/기호 (선택)",
            value="",
            placeholder="예: ㄱ, ★, [회사]",
            help="예) 'ㄱ' 입력 → '홍길동' → 'ㄱ홍길동' 으로 저장됩니다.",
        )
    with oc3:
        skip_no_phone = st.checkbox("휴대폰이 없으면 제외", value=True)

    # 데이터 수집
    def col_idx(letter):
        return column_index_from_string(letter) if letter else None

    iB = col_idx(L_BRANCH); iM = col_idx(L_MANAGER)
    iA = col_idx(L_AGENCY); iS = col_idx(L_SUB)
    iN = col_idx(L_NAME); iP = col_idx(L_PHONE); iNote = col_idx(L_NOTE)

    rows_data = []
    for row in ws.iter_rows(min_row=int(data_start_row), max_row=max_row, values_only=True):
        def g(idx):
            if idx is None or idx > len(row): return ""
            return cell_value(row[idx - 1])
        branch = g(iB); manager = g(iM)
        agency = g(iA); sub = g(iS)
        name = g(iN); phone_raw = g(iP); note = g(iNote)
        if not branch or not name:
            continue
        clean_phone = normalize_phone(phone_raw)
        if skip_no_phone and not clean_phone:
            continue
        if name_prefix:
            name = f"{name_prefix}{name}"
        org = sub or agency
        final_note = note
        if extra_note:
            final_note = (note + " " + extra_note).strip() if note else extra_note
        rows_data.append({
            "지점": branch,
            "매니저": manager or "_매니저미지정",
            "대리점": agency, "지사": sub,
            "이름": name, "휴대폰": clean_phone or phone_raw,
            "_clean_phone": clean_phone,
            "ORG": org, "메모": final_note,
        })

    if not rows_data:
        st.warning("유효한 데이터가 없습니다. 컬럼 매핑과 데이터 시작 행을 확인해주세요.")
        st.stop()

    # Step 5: 미리보기 & 통계
    st.markdown("---")
    st.markdown(
        f'<span class="step-badge">5</span> **미리보기 & 통계** &nbsp; <span class="count-badge">👤 {len(rows_data)}건</span>',
        unsafe_allow_html=True,
    )

    preview_df = pd.DataFrame(rows_data)[["지점", "매니저", "지사", "대리점", "이름", "휴대폰", "메모"]]
    st.dataframe(preview_df.head(15), use_container_width=True, hide_index=True)
    if len(preview_df) > 15:
        st.caption(f"... 외 {len(preview_df) - 15}건 더")

    summary = preview_df.groupby(["지점", "매니저"]).size().reset_index(name="인원수")
    st.caption("**지점 · 매니저별 인원수**")
    st.dataframe(summary, use_container_width=True, hide_index=True, height=260)

    # Step 6: ZIP 생성
    st.markdown("---")
    st.markdown('<span class="step-badge">6</span> **VCF ZIP 생성**', unsafe_allow_html=True)

    groups = defaultdict(lambda: defaultdict(list))
    for r in rows_data:
        groups[r["지점"]][r["매니저"]].append(r)

    zip_buf = io.BytesIO()
    file_count = 0
    contact_count = 0
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for branch in sorted(groups.keys()):
            for manager in sorted(groups[branch].keys()):
                rows = groups[branch][manager]
                cnt = len(rows)
                fname = f"{sanitize_filename(manager)}_{cnt}.vcf"
                arcname = f"{sanitize_filename(branch)}/{fname}"
                body = ""  # iOS 호환: BOM 없이 BEGIN:VCARD로 바로 시작
                for r in rows:
                    fn = f"{r['이름']} ({r['ORG']})" if r["ORG"] else r["이름"]
                    body += to_vcard(fn, r["_clean_phone"] or r["휴대폰"], org=r["ORG"], note=r["메모"])
                zf.writestr(arcname, body.encode("utf-8"))
                file_count += 1
                contact_count += cnt
    zip_buf.seek(0)

    st.download_button(
        f"📦 ZIP 다운로드 ({len(groups)}지점 · {file_count}파일 · {contact_count}건)",
        data=zip_buf.getvalue(),
        file_name=f"연락처_매니저별_{contact_count}건.zip",
        mime="application/zip",
        use_container_width=True,
        type="primary",
    )

    st.markdown(
        f'<div class="success-box">✅ ZIP 안에 <b>지점 폴더 / 매니저명_인원수.vcf</b> 구조로 들어있습니다.<br>'
        f'· 지점: {len(groups)}개 &nbsp; · VCF 파일: {file_count}개 &nbsp; · 연락처: {contact_count}건</div>',
        unsafe_allow_html=True,
    )

# 사용 가이드
st.markdown("---")
st.markdown(
    """<div class="guide-box">
💡 <b>사용 방법</b><br>
<b>1.</b> 엑셀 파일 업로드 (양식 그대로 OK)<br>
<b>2.</b> 헤더 행 / 데이터 시작 행 입력<br>
<b>3.</b> 각 항목을 <b>실제 엑셀 열 문자(A, B, ..., AF)</b>로 선택<br>
<b>4.</b> 미리보기 → ZIP 다운로드<br><br>
📌 <b>예시 (사용인검색목록 양식):</b><br>
&nbsp;&nbsp;지점=<b>G</b>, 매니저=<b>I</b>, 대리점=<b>K</b>, 지사=<b>M</b>, 이름=<b>P</b>, 휴대폰=<b>AF</b><br>

📌 <b>vCard 형식:</b> FN(이름+지사), TEL, ORG(지사), NOTE(메모) — vCard 3.0
</div>""",
    unsafe_allow_html=True,
)
